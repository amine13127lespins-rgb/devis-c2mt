#!/usr/bin/env python3
"""
Script de prospection - récupère des établissements (restaurants, bars, ...)
via l'API Google Places (New) pour une liste de villes/catégories, et
génère un fichier Excel de prospection prêt à l'emploi.

Voir scripts/prospection/README.md pour le mode d'emploi complet.
"""

import re
import time
import unicodedata

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIGURATION - à modifier avant de lancer le script
# ============================================================

# Colle ta clé API Google Places (New) ici, entre les guillemets.
API_KEY = "VOTRE_CLE_API_ICI"

# Liste des villes dans lesquelles chercher (ajoute/retire des lignes librement).
VILLES = [
    "Nice",
    "Cannes",
    "Antibes",
    "Marseille",
    "Aix-en-Provence",
]

# Liste des catégories d'établissements recherchées.
CATEGORIES = [
    "restaurant",
    "pizzeria",
    "bar",
    "brasserie",
    "café",
    "boulangerie",
    "food truck",
    "glacier",
    "crêperie",
]

# Nom du fichier Excel généré (créé dans le même dossier que ce script).
NOM_FICHIER_SORTIE = "prospects_restaurants_paca.xlsx"

# ============================================================
# CONSTANTES TECHNIQUES - normalement pas besoin d'y toucher
# ============================================================

SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"
FIELD_MASK = (
    "places.displayName,places.formattedAddress,"
    "places.nationalPhoneNumber,places.internationalPhoneNumber,"
    "places.websiteUri,places.rating,places.userRatingCount,"
    "nextPageToken"
)
PAGE_SIZE = 20  # maximum autorisé par l'API pour la recherche texte
MAX_PAGES = 3  # Google limite à ~60 résultats (3 pages de 20) par recherche
DELAI_ENTRE_PAGES = 2  # secondes, le nextPageToken n'est actif qu'après ce délai
DELAI_ENTRE_REQUETES = 0.3  # secondes, pour rester tranquille avec les quotas

STATUTS = ["À contacter", "Contacté", "Relancé", "Intéressé", "Vendu", "Refusé"]
OUI_NON = ["Oui", "Non"]

COLONNES = [
    "N°", "Établissement", "Type", "Ville", "Téléphone", "Instagram",
    "Statut", "Canal", "Date contact", "Date relance",
    "Vente 30€", "Abo 15€/mois", "Notes",
]


def normaliser(texte):
    """Enlève les accents et met en minuscule, pour comparer deux textes lors de la déduplication."""
    if not texte:
        return ""
    texte = unicodedata.normalize("NFKD", texte).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", texte).strip().lower()


def rechercher_etablissements(categorie, ville):
    """Interroge l'API Google Places pour une catégorie et une ville, en suivant la pagination."""
    resultats = []
    page_token = None

    for _ in range(MAX_PAGES):
        corps_requete = {
            "textQuery": f"{categorie} à {ville}, France",
            "languageCode": "fr",
            "pageSize": PAGE_SIZE,
        }
        if page_token:
            corps_requete["pageToken"] = page_token

        entetes = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": API_KEY,
            "X-Goog-FieldMask": FIELD_MASK,
        }

        reponse = requests.post(SEARCH_URL, json=corps_requete, headers=entetes, timeout=30)

        if reponse.status_code != 200:
            print(f"  ! Erreur API ({reponse.status_code}) pour '{categorie} à {ville}' : {reponse.text[:300]}")
            break

        donnees = reponse.json()
        resultats.extend(donnees.get("places", []))

        page_token = donnees.get("nextPageToken")
        if not page_token:
            break

        time.sleep(DELAI_ENTRE_PAGES)

    return resultats


def construire_ligne(place, categorie, ville):
    nom = place.get("displayName", {}).get("text", "").strip()
    adresse = place.get("formattedAddress", "")
    telephone = place.get("nationalPhoneNumber") or place.get("internationalPhoneNumber") or ""
    site = place.get("websiteUri", "")
    note = place.get("rating", "")
    nb_avis = place.get("userRatingCount", "")

    morceaux = [
        f"Adresse : {adresse}" if adresse else None,
        f"Site : {site}" if site else None,
        f"Note Google : {note}/5 ({nb_avis} avis)" if note != "" else None,
    ]
    notes = " | ".join(m for m in morceaux if m)

    return {
        "nom": nom,
        "type": categorie,
        "ville": ville,
        "telephone": telephone,
        "notes": notes,
    }


def collecter_prospects():
    prospects = {}  # clé de dédup (nom+ville normalisés) -> ligne

    for ville in VILLES:
        for categorie in CATEGORIES:
            print(f"Recherche : {categorie} à {ville}...")
            places = rechercher_etablissements(categorie, ville)
            print(f"  -> {len(places)} résultat(s)")

            for place in places:
                ligne = construire_ligne(place, categorie, ville)
                if not ligne["nom"]:
                    continue
                cle = (normaliser(ligne["nom"]), normaliser(ville))
                if cle not in prospects:
                    prospects[cle] = ligne

            time.sleep(DELAI_ENTRE_REQUETES)

    return list(prospects.values())


def generer_excel(prospects):
    classeur = Workbook()
    feuille = classeur.active
    feuille.title = "Prospects"

    feuille.append(COLONNES)
    police_entete = Font(bold=True, color="FFFFFF")
    fond_entete = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    for col in range(1, len(COLONNES) + 1):
        cellule = feuille.cell(row=1, column=col)
        cellule.font = police_entete
        cellule.fill = fond_entete
        cellule.alignment = Alignment(horizontal="center", vertical="center")

    for i, p in enumerate(prospects, start=1):
        feuille.append([
            i,
            p["nom"],
            p["type"],
            p["ville"],
            p["telephone"],
            "",              # Instagram (à compléter à la main)
            "À contacter",   # Statut
            "",              # Canal
            "",              # Date contact
            "",              # Date relance
            "",              # Vente 30€
            "",              # Abo 15€/mois
            p["notes"],
        ])

    nb_lignes = max(len(prospects) + 1, 2)

    largeurs = [5, 30, 16, 16, 16, 16, 14, 12, 14, 14, 12, 14, 65]
    for idx, largeur in enumerate(largeurs, start=1):
        feuille.column_dimensions[get_column_letter(idx)].width = largeur

    def ajouter_liste_deroulante(colonne_lettre, valeurs):
        dv = DataValidation(
            type="list",
            formula1='"{}"'.format(",".join(valeurs)),
            allow_blank=True,
            showDropDown=False,  # False = affiche bien la flèche de liste déroulante
        )
        feuille.add_data_validation(dv)
        dv.add(f"{colonne_lettre}2:{colonne_lettre}{nb_lignes}")

    ajouter_liste_deroulante("G", STATUTS)     # Statut
    ajouter_liste_deroulante("C", CATEGORIES)  # Type
    ajouter_liste_deroulante("K", OUI_NON)     # Vente 30€
    ajouter_liste_deroulante("L", OUI_NON)     # Abo 15€/mois

    feuille.freeze_panes = "A2"
    classeur.save(NOM_FICHIER_SORTIE)
    print(f"\nFichier généré : {NOM_FICHIER_SORTIE} ({len(prospects)} prospects uniques)")


def main():
    if not API_KEY or API_KEY == "VOTRE_CLE_API_ICI":
        print("Merci de renseigner votre clé API Google dans la variable API_KEY en haut du script.")
        return

    prospects = collecter_prospects()
    if not prospects:
        print("Aucun établissement trouvé. Vérifiez vos villes/catégories ou votre clé API.")
        return

    generer_excel(prospects)


if __name__ == "__main__":
    main()
